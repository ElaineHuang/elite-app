#-*-coding:utf-8 -*-
import os
import openpyxl
import operator
import time
import glob
import datetime
from datetime import timedelta
"""
SMT production line
data preprocessing
"""
tagname_to_machine_name = {
	'04': 'M023',
	'05': 'M024',
	'06': 'M025',
	'10': 'M026'
}
def sort_table(table, cols):
    """ sort a table by multiple columns
        table: a list of lists (or tuple of tuples) where each inner list 
               represents a row
        cols:  a list (or tuple) specifying the column numbers to sort by
               e.g. (1,0) would sort by column 1, then by column 0
    """
    for col in reversed(cols):
        table = sorted(table, key=operator.itemgetter(col))
    return table

def calculate_light():
	## Adjusting sdTimeLimit if you want to change the low bundary
	sdTimeLimit = 180
	# raw_data_path = os.path.join(os.path.abspath(os.path.dirname('data')), 'data/light_data/')
	# useful_data_path = os.path.join(os.path.abspath(os.path.dirname('data')), 'data/useful_light_data/')		
	raw_data_path = os.path.join(os.path.abspath(os.path.dirname('../data')), 'data/light_data/')
	useful_data_path = os.path.join(os.path.abspath(os.path.dirname('../data')), 'data/useful_light_data/')	
	files_list = sorted(glob.glob(raw_data_path + "*.xlsx"))	
	SpecList = ['04', '05', '06', '10']


	files = [f for f in files_list if os.path.isfile(os.path.join(raw_data_path, f))]
	try:
		print files
		# SMT_path = 'D:\Documents\IoT\preprocessing\ColoredLight'
		for f in files:
			for index in range(len(SpecList)):
				fileName = os.path.splitext(f)[0].split('/')[-1]
				unsortedTable = []
				perDataSize=[0]*len(SpecList)
				workbook = openpyxl.load_workbook(filename=os.path.join(raw_data_path, f), read_only=True)

				#輸出的檔名
				# LightRec = open(useful_data_path + fileName + tagname_to_machine_name[SpecList[index]]+"_light.csv","w")
				LightRec = open(useful_data_path + tagname_to_machine_name[SpecList[index]]+"_light.csv","w")			
				LightRec.write("Situation, Logtime, Duration\n")			
				for sheetName in workbook.get_sheet_names():
					sheet = workbook.get_sheet_by_name(sheetName)
					if sheet.cell(row = 1, column=1).value == None:
						break
					# if sheetName == 'simulation':
					else:
						Datasize = sheet.max_row-1
						for row in sheet.iter_rows(min_row=2, min_col=1, max_col=5, max_row=Datasize+1):
							model=row[0].value.strip(' ')
							if not isinstance(row[2].value, datetime.time):
								if model[-2:] == SpecList[index]:
									minnnn = timedelta(hours=row[2].value*24)
									# dddddd = timedelta(milliseconds=)
									tttttt = row[1].value + minnnn
									unsortedTable.append([model[2], tttttt, row[4].value])
									perDataSize[index] += 1						
							else:
								if model[-2:] == SpecList[index]:
									unsortedTable.append([model[2], datetime.datetime.combine(row[1].value.date(), row[2].value), str(row[4].value)])
									perDataSize[index] += 1
				sortedTable = sort_table(unsortedTable, (1,2))

				if len(sortedTable) > 0:
					temp = 0
					coloredLight = {'G' : "0", 'R' : "0", 'Y' : "0"}
					lastTime = sortedTable[0][1]
					nowTime = sortedTable[0][1]
					listEnd = 0
					while temp < perDataSize[index]:
						while nowTime == sortedTable[temp][1]:
							coloredLight[sortedTable[temp][0]] = sortedTable[temp][2]
							temp += 1
							if temp >= perDataSize[index]:
								listEnd = 1
								break
						
						if listEnd == 1:
							break
						
						lastTime = nowTime
						nowTime = sortedTable[temp][1]
						
						if (isinstance(coloredLight['G'], long)):
							coloredLight['G'] = str(coloredLight['G'])
						if (isinstance(coloredLight['R'], long)):
							coloredLight['R'] = str(coloredLight['R'])
						if (isinstance(coloredLight['Y'], long)):
							coloredLight['Y'] = str(coloredLight['Y'])														
						if not((coloredLight['G']=="0") and (coloredLight['R']=="1") and (coloredLight['Y']=="0")):
							continue

						duration = int((nowTime-lastTime).total_seconds())
						if duration>=sdTimeLimit:
							line = ["shutdown", lastTime.strftime("%Y/%m/%d %H:%M:%S"), str(duration)]
							# print sheetName, line
							LightRec.write(','.join(line)+"\n")

				LightRec.close()
	except Exception as e:
		print e
		raise e
if __name__ == '__main__':
	calculate_light()