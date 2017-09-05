#-*-coding:utf-8 -*-
import os
import glob
import openpyxl
import operator
import datetime
"""
SMT production line
data preprocessing
"""

def sort_table(table, cols):
    """ sort a table by multiple columns
        table: a list of lists (or tuple of tuples) where each inner list 
               represents a row
        cols:  a list (or tuple) specifying the column numbers to sort by
               e.g. (1,0) would sort by column 1, then by column 0
    """
    for col in reversed(cols):
        table = sorted(table, key=operator.itemgetter(col))
    return table


def processing_to_csv():
	SpecList = ['E620002L', 'E620002R', 'E620003L', 'E620003R', 'E551010L', 'E551010R']
	raw_data_path = os.path.join(os.path.abspath(os.path.dirname('data')), 'data/raw_data/')
	useful_data_path = os.path.join(os.path.abspath(os.path.dirname('data')), 'data/useful_data/')
	files_list = sorted(glob.glob(raw_data_path + "*.xlsx"))
	files = [f for f in files_list if os.path.isfile(os.path.join(raw_data_path, f))]
	print files
	for f in files:
		fileName = os.path.splitext(f)[0].split('/')[-1]
		ErrRec = open(useful_data_path + fileName+"_file.csv","w")
		print ("-----------processing" + fileName + "_file------------")
		ErrRec.write("StartTime,ErrCode,Repetition\n")
		unsortedTable = []
		workbook = openpyxl.load_workbook(filename=os.path.join(raw_data_path, f), read_only=True)
		
		sheet = workbook.get_sheet_by_name('EQErrRec_'+ fileName.split('_')[0])
		Datasize = sheet.max_row-1
		for row in sheet.iter_rows(min_row=2, min_col=2, max_col=4, max_row=Datasize+1):
			unsortedTable.append([row[0].value, row[1].value])
		sortedTable = sort_table(unsortedTable, (0,1))
		del unsortedTable
		
		TimeDiffTable=[]
		standardTime = sortedTable[0][0]
		TimeDiffTable.append([0,sortedTable[0][1]])
		for row in range(1,len(sortedTable)):
			if not isinstance(sortedTable[row][0], datetime.datetime):
				time_cc = (datetime.datetime.fromtimestamp(sortedTable[row][0]/1000) - datetime.datetime.fromtimestamp(standardTime/1000)).total_seconds()
			else:
				time_cc = (sortedTable[row][0] - standardTime).total_seconds()
			TimeDiffTable.append([time_cc,sortedTable[row][1]])
		
		index = 0
		for row in TimeDiffTable:
			dirty = 0
			baseerror = str(row[1])
			if baseerror == '0':
				index += 1
				continue
			basetime = row[0]		
			temp = index+1
			if (temp >= Datasize):
				dirty = 1
			count = 1
			if (baseerror in SpecList)  and (dirty == 0):
				while (TimeDiffTable[temp][0]-basetime) <= 120:
					if (str(TimeDiffTable[temp][1]) == baseerror):
						count += 1
						TimeDiffTable[temp]=[0,'0']
					temp += 1
					if(temp >= Datasize):
						break
			elif (baseerror not in SpecList) and (dirty == 0):
				while (TimeDiffTable[temp][0]-basetime) <= 300:
					if (str(TimeDiffTable[temp][1]) == baseerror):
						count += 1
						TimeDiffTable[temp]=[0,'0']
					temp += 1
					if(temp >= Datasize):
						break
			TimeDiffTable[index][0] = count
			index += 1
		
		index = 0
		for row in TimeDiffTable:
			baseerror = str(row[1])
			if baseerror == '0':
				index += 1;
				continue
			if not isinstance(sortedTable[index][0], datetime.datetime):
				time_rd = datetime.datetime.fromtimestamp(sortedTable[index][0]/1000)
			else:
				time_rd = sortedTable[index][0]
			line = [time_rd.strftime("%Y/%m/%d %H:%M:%S"), baseerror, str(row[0])]
			ErrRec.write(','.join(line)+"\n")
			index += 1;
		ErrRec.close()
		print ("============== done" + fileName + "_file ===========")

def processing_to_db(db, collect_name):

	machine_all_data = list(db[collect_name].find())
	entries = list(machine_all_data[:])
	print (entries)
	# for data in machine_all_data:
	# 	print (data)

